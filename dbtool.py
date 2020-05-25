from openpyxl import load_workbook
import datetime
import sqlite3
import sys
import os

def format_value(col_value):
    if type(col_value) == datetime.datetime:
        return f"'{str(col_value.date())}'"
    elif type(col_value) == str:
        return f"'{col_value}'"
    else:
        return col_value

def retrieve_path_and_dbname(arguments):
    try:
        supported_formats = tuple({'xlsx', 'xlsm', 'xlxt', 'xltm'})

        if (len(arguments) < 2):
            raise Exception('Path to file not specified')
        elif (len(arguments) > 2):
            raise Exception('Too many input arguments')

        path = arguments[1]
        
        if(not os.path.isfile(path)):
            raise Exception('Path does not exists or is not of a file')
        
        filename = os.path.split(path)[1]
        dbname, extension = filename.split(".")

        if (extension not in supported_formats):
            raise Exception(f'''Format not supported '{extension}'.
Supported formats are: .xlsx,.xlsm,.xltx,.xltm''')

        return (path, dbname + '.db')

    except Exception as e:
        print(e)
        sys.exit(1)

def get_datatypes(iter):
    datatypes = list()
   
    for item in next(sheet_iter):  
        type_of_item = type(item)
        
        if type_of_item == int:
            datatypes.append('INTEGER')
        elif type_of_item == float:
            datatypes.append('REAL')
        else:
            datatypes.append('STRING')
    
    return datatypes

path, db_name = retrieve_path_and_dbname(sys.argv)

workbook = load_workbook(path, read_only=True)
table_name = workbook.sheetnames[0].upper()
sheet = workbook.active
sheet_iter = sheet.values

column_names = tuple(next(sheet_iter))
column_names = tuple(map(lambda elem: elem.upper().replace(" ", "_"), column_names))

datatypes = get_datatypes(sheet_iter)

create_table_query = f'create table {table_name} ('
joined_colName_and_types = ', '.join(f"{i} {j}" for (i,j) in zip(column_names, datatypes))
create_table_query += joined_colName_and_types + ');'

conn = sqlite3.connect(db_name)
cur = conn.cursor()

cur.execute(create_table_query)

sheet_iter = sheet.values
next(sheet_iter)
try:
    while(True):
        row = map(lambda x: format_value(x), next(sheet_iter))
        insert_query = f'insert into {table_name} values('
        joined_rows  = ', '.join(f'{i}' for i in row)
        insert_query += joined_rows + ');'
        cur.execute(insert_query)

except StopIteration:
    conn.commit()
finally:
    conn.close()
    workbook.close()
